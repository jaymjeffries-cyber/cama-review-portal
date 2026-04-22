import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime
from openpyxl import load_workbook

# Page configuration
st.set_page_config(
    page_title="MLS vs CAMA Data Comparison",
    page_icon="📊",
    layout="wide"
)

# --- Configuration ---
UNIQUE_ID_COLUMN = {'mls_col': 'Parcel Number', 'cama_col': 'PARID'}

COLUMNS_TO_COMPARE = [
    {'mls_col': 'Above Grade Finished Area', 'cama_col': 'SFLA'},
    {'mls_col': 'Bedrooms Total', 'cama_col': 'RMBED'},
    {'mls_col': 'Bathrooms Full', 'cama_col': 'FIXBATH'},
    {'mls_col': 'Bathrooms Half', 'cama_col': 'FIXHALF'},
]

COLUMNS_TO_COMPARE_SUM = [
    {'mls_col': 'Below Grade Finished Area', 'cama_cols': ['RECROMAREA', 'FINBSMTAREA', 'UFEATAREA']}
]

COLUMNS_TO_COMPARE_CATEGORICAL = [
    {
        'mls_col': 'Cooling',
        'cama_col': 'HEAT',
        'mls_check_contains': 'Central Air',
        'cama_expected_if_true': 1,
        'cama_expected_if_false': 0,
        'case_sensitive': False
    }
]

NUMERIC_TOLERANCE = 0.01
SKIP_ZERO_VALUES = True

ADDRESS_COLUMNS = {
    'address': 'Address',
    'city': 'City',
    'state': 'State or Province',
    'zip': 'Postal Code'
}

ZILLOW_URL_BASE = "https://www.zillow.com/homes/"

# --- TAXDIST to Zone Mapping ---
TAXDIST_TO_ZONE = {
    '00010': 'North East',
    '00020': 'Canton',
    '00025': 'Canton',
    '00030': 'Canton',
    '00035': 'Canton',
    '00040': 'Western',
    '00050': 'Western',
    '00060': 'Western',
    '00065': 'Western',
    '00070': 'Southern',
    '00080': 'Southern',
    '00090': 'Southern',
    '00100': 'Southern',
    '00110': 'Southern',
    '00112': 'Southern',
    '00115': 'Southern',
    '00120': 'Southern',
    '00130': 'North Western',
    '00140': 'North Western',
    '00150': 'North Western',
    '00160': 'Northern',
    '00170': 'Northern',
    '00180': 'Northern',
    '00190': 'Northern',
    '00200': 'Northern',
    '00210': 'North Western',
    '00220': 'North Western',
    '00230': 'North Western',
    '00240': 'North Western',
    '00245': 'North Western',
    '00250': 'North East',
    '00260': 'North East',
    '00270': 'North East',
    '00280': 'North East',
    '00290': 'North East',
    '00300': 'North East',
    '00305': 'North East',
    '00310': 'North East',
    '00320': 'North East',
    '00330': 'North East',
    '00340': 'Southern',
    '00345': 'Southern',
    '00350': 'Southern',
    '00355': 'Southern',
    '00360': 'Southern',
    '00370': 'Southern',
    '00380': 'Southern',
    '00390': 'Southern',
    '00400': 'Western',
    '00410': 'Western',
    '00415': 'Southern',
    '00420': 'Western',
    '00430': 'Western',
    '00440': 'Western',
    '00445': 'Southern',
    '00450': 'Southern',
    '00460': 'Southern',
    '00470': 'Southern',
    '00480': 'Southern',
    '00490': 'Northern',
    '00520': 'Northern',
    '00530': 'North Western',
    '00535': 'North Western',
    '00545': 'North Western',
    '00555': 'North Western',
    '00560': 'Northern',
    '00565': 'Canton',
    '00570': 'Southern',
    '00580': 'Southern',
    '00590': 'Southern',
    '00600': 'Southern',
    '00610': 'Southern',
    '00620': 'Southern',
    '00630': 'Southern',
    '00640': 'Southern',
    '00660': 'Southern',
    '00670': 'Southern',
    '00680': 'Southern',
    '00690': 'Western',
    '00700': 'Western',
    '00710': 'Western',
    '00715': 'North East',
    '00720': 'North East',
    '00725': 'North East',
    '00730': 'North East',
    '00740': 'North East',
    '00750': 'North East',
}

def lookup_zone(taxdist):
    """Normalize TAXDIST to 5-digit zero-padded string and return Zone name."""
    if pd.isna(taxdist) or str(taxdist).strip() == '':
        return 'Unknown'
    try:
        key = str(int(float(str(taxdist)))).zfill(5)
    except (ValueError, TypeError):
        key = str(taxdist).strip().zfill(5)
    return TAXDIST_TO_ZONE.get(key, 'Unknown')

# --- Helper Functions ---

def format_zillow_url(address, city, state, zip_code):
    if pd.isna(address) or pd.isna(city) or pd.isna(zip_code):
        return None
    import re
    address_clean = str(address).strip()
    city_clean = str(city).strip()
    zip_clean = str(zip_code).strip().split('-')[0]
    address_clean = re.sub(r'\s+(Apt|Unit|#|Suite)\s*[\w-]*$', '', address_clean, flags=re.IGNORECASE)
    address_formatted = re.sub(r'[^\w\s-]', '', address_clean)
    address_formatted = re.sub(r'\s+', '-', address_formatted)
    city_formatted = re.sub(r'[^\w\s-]', '', city_clean)
    city_formatted = re.sub(r'\s+', '-', city_formatted)
    url_slug = f"{address_formatted}-{city_formatted}-OH-{zip_clean}_rb"
    return f"{ZILLOW_URL_BASE}{url_slug}/"

def load_mls_file(uploaded_file):
    """
    Load MLS Excel file and auto-fill blank cells with 0.
    Replaces the manual CTRL+H blank-to-zero step done each week.
    """
    df = pd.read_excel(uploaded_file)
    blank_count = int(df.isnull().sum().sum())
    df = df.fillna(0)
    df = df.replace('', 0)
    return df, blank_count

def create_mls_search_batches_excel(df_missing_in_mls, batch_size=35):
    """
    Create an Excel file with missing_in_MLS parcels formatted for MLS search.
    - Parcel numbers formatted to 8 digits with leading zeros
    - Split into tabs of batch_size parcels each (max 35 per MLS search)
    - Each tab has a ready-to-paste column
    """
    output = io.BytesIO()

    parcel_ids = df_missing_in_mls['Parcel_ID'].tolist()
    formatted_parcels = []
    for p in parcel_ids:
        try:
            formatted = str(int(float(str(p)))).zfill(8)
        except (ValueError, TypeError):
            formatted = str(p).zfill(8)
        formatted_parcels.append(formatted)

    total = len(formatted_parcels)
    num_batches = max(1, (total + batch_size - 1) // batch_size)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary tab
        summary_data = {
            'Total Parcels': [total],
            'Batch Size': [batch_size],
            'Number of Batches': [num_batches],
            'Generated': [datetime.now().strftime('%m/%d/%Y %H:%M')]
        }
        pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name='Summary')
        ws_sum = writer.sheets['Summary']
        ws_sum.column_dimensions['A'].width = 16
        ws_sum.column_dimensions['B'].width = 14
        ws_sum.column_dimensions['C'].width = 18
        ws_sum.column_dimensions['D'].width = 20

        # One tab per batch — comma-separated string for direct MLS paste
        for i in range(num_batches):
            start = i * batch_size
            end = min(start + batch_size, total)
            batch = formatted_parcels[start:end]
            comma_str = ','.join(batch)

            sheet_name = f'Batch {i+1} ({start+1}-{end})'
            pd.DataFrame([[]]).to_excel(writer, index=False, header=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            ws['A1'] = f'Batch {i+1} — {len(batch)} parcels — copy cell A2 and paste into MLS parcel search'
            ws['A2'] = comma_str
            ws.column_dimensions['A'].width = 120
            from openpyxl.styles import Font, PatternFill, Alignment
            ws['A1'].font = Font(bold=True, color='FFFFFF', size=11)
            ws['A1'].fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            ws['A2'].alignment = Alignment(wrap_text=False)
            ws.row_dimensions[2].height = 20

    output.seek(0)
    return output.getvalue(), num_batches

def values_equal(val1, val2):
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')
        if pd.isna(num1) and pd.isna(num2):
            return True
        elif pd.isna(num1) != pd.isna(num2):
            return False
        else:
            return np.isclose(num1, num2, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except (ValueError, TypeError):
        str1 = str(val1).strip().lower() if pd.notna(val1) else ''
        str2 = str(val2).strip().lower() if pd.notna(val2) else ''
        return str1 == str2

def categorical_match(mls_val, cama_val, mapping):
    check_text = mapping.get('mls_check_contains', '')
    expected_if_true = mapping.get('cama_expected_if_true')
    expected_if_false = mapping.get('cama_expected_if_false')
    case_sensitive = mapping.get('case_sensitive', False)
    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''
    if not case_sensitive:
        mls_str = mls_str.lower()
        check_text = check_text.lower()
    text_found = check_text in mls_str
    expected_cama = expected_if_true if text_found else expected_if_false
    try:
        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
        expected_numeric = pd.to_numeric(expected_cama, errors='coerce')
        if pd.isna(cama_numeric) and pd.isna(expected_numeric):
            return True
        elif pd.isna(cama_numeric) or pd.isna(expected_numeric):
            return False
        else:
            return np.isclose(cama_numeric, expected_numeric, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except:
        return str(cama_val).strip().lower() == str(expected_cama).strip().lower()

def calculate_difference(val1, val2):
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')
        if pd.isna(num1) or pd.isna(num2):
            return "N/A"
        diff = num1 - num2
        return f"{diff:,.2f}"
    except (ValueError, TypeError):
        return "Text difference"

def generate_mass_update_files(combined_df, user_initials, user_full_name):
    required_cols = ['Parcel_ID', 'SALEKEY']
    missing_cols = [col for col in required_cols if col not in combined_df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns: {missing_cols}")

    listing_col = 'Listing_Number' if 'Listing_Number' in combined_df.columns else ('Listing #' if 'Listing #' in combined_df.columns else None)
    unique_df = combined_df.drop_duplicates(subset=['Parcel_ID'], keep='first').copy()

    saletab_rows = []
    entrance_rows = []

    for idx, row in unique_df.iterrows():
        main_parcel = row['Parcel_ID']
        salekey_str = str(row.get('SALEKEY', '')).strip()
        additional_parcels_str = str(row.get('ADDITIONAL_PARCELS', '')).strip()

        listing_number = 0
        if listing_col:
            listing_str = str(row.get(listing_col, '')).strip()
            if listing_str and listing_str != 'nan':
                first_listing = listing_str.rstrip(',').split(',')[0].strip()
                try:
                    listing_number = int(float(first_listing))
                except (ValueError, TypeError):
                    listing_number = 0

        salekeys = []
        if salekey_str and salekey_str != 'nan':
            salekeys = [s.strip() for s in salekey_str.rstrip(',').split(',') if s.strip()]

        additional_parcels = []
        if additional_parcels_str and additional_parcels_str != 'nan':
            additional_parcels = [p.strip() for p in additional_parcels_str.split(',') if p.strip()]

        all_parcels = [main_parcel] + additional_parcels

        for i, parcel in enumerate(all_parcels):
            if i < len(salekeys):
                try:
                    salekey_int = int(salekeys[i])
                    parcel_int = int(parcel)
                    saletab_rows.append({
                        'PARID': parcel_int,
                        'SALEKEY': salekey_int,
                        'USER11': listing_number,
                        'SOURCE': 0,
                        'SALEVAL': 0,
                        'USER1': user_initials,
                        'USER2': pd.Timestamp.now().strftime('%Y-%m-%d')
                    })
                    entrance_rows.append({
                        'Change Type': 'existing',
                        'appraiser': user_full_name,
                        'parcelnum': parcel_int,
                        'comment': '',
                        'Review Status': 'Reviewed',
                        'Determination': '',
                        'Est. Value Change': '',
                        'Last Changed Date/Time': pd.Timestamp.now().strftime('%m/%d/%Y %H:%M'),
                        'Last Changed By': user_full_name
                    })
                except (ValueError, TypeError):
                    continue

    saletab_df = pd.DataFrame(saletab_rows)
    if not saletab_df.empty:
        saletab_df = saletab_df.drop_duplicates(subset=['SALEKEY'], keep='first')
        saletab_df = saletab_df.sort_values('SALEKEY').reset_index(drop=True)

    saletab_buffer = io.BytesIO()
    with pd.ExcelWriter(saletab_buffer, engine='openpyxl') as writer:
        saletab_df.to_excel(writer, index=False, sheet_name='Sheet1')
    saletab_buffer.seek(0)

    entrance_df = pd.DataFrame(entrance_rows)
    entrance_csv = entrance_df.to_csv(index=False)

    return saletab_buffer.getvalue(), entrance_csv

def compare_data_enhanced(df_mls, df_cama, unique_id_col, cols_to_compare_mapping,
                         cols_to_compare_sum=None, cols_to_compare_categorical=None,
                         window_id=None):
    if df_mls is None or df_cama is None:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    mls_id_col_name = unique_id_col.get('mls_col')
    cama_id_col_name = unique_id_col.get('cama_col')

    if mls_id_col_name not in df_mls.columns:
        st.error(f"Column '{mls_id_col_name}' not found in MLS data")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    if cama_id_col_name not in df_cama.columns:
        st.error(f"Column '{cama_id_col_name}' not found in CAMA data")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df_mls_renamed = df_mls.copy()
    df_mls_renamed = df_mls_renamed.rename(columns={mls_id_col_name: cama_id_col_name})

    matched_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='inner')
    merged_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='outer', indicator=True)

    missing_in_cama = []
    missing_in_mls = []
    value_mismatches = []
    perfect_matches = []

    if window_id:
        parcel_url_template = f"https://iasworld.starkcountyohio.gov/iasworld/Maintain/Transact.aspx?txtMaskedPin={{parcel_id}}&selYear=&userYear=&selJur=&chkShowHistory=False&chkShowChanges=&chkShowDeactivated=&PinValue={{parcel_id}}&pin=&trans_key=&windowId={window_id}&submitFlag=true&TransPopUp=&ACflag=False&ACflag2=False"
    else:
        parcel_url_template = None

    for index, row in merged_df.iterrows():
        record_id = row.get(cama_id_col_name)
        merge_status = row.get('_merge')

        if merge_status == 'left_only':
            listing_num = row.get('Listing #', '')
            closed_date = row.get('Closed Date', '')
            missing_in_cama.append({
                'Parcel_ID': record_id,
                'Listing_Number': listing_num,
                'Closed_Date': closed_date
            })

        elif merge_status == 'right_only':
            missing_in_mls.append({'Parcel_ID': record_id})

        elif merge_status == 'both':
            listing_num = row.get('Listing #', '')
            salekey = row.get('SALEKEY', '')
            nopar = row.get('NOPAR', '')
            additional_parcels = row.get('ADDITIONAL_PARCELS', '')

            address = row.get(ADDRESS_COLUMNS.get('address', 'Address'), '')
            city = row.get(ADDRESS_COLUMNS.get('city', 'City'), '')
            state = row.get(ADDRESS_COLUMNS.get('state', 'State or Province'), '')
            zip_code = row.get(ADDRESS_COLUMNS.get('zip', 'Postal Code'), '')

            # Extra context columns for value_mismatches and perfect_matches
            yrblt = row.get('YRBLT', '')
            effyr = row.get('EFFYR', '')
            grade = row.get('GRADE', '')
            cdu = row.get('CDU', '')
            public_remarks = row.get('Public Remarks', '')
            taxdist = row.get('TAXDIST', '')
            zone = lookup_zone(taxdist)
            fixbath = row.get('FIXBATH', '')
            fixhalf = row.get('FIXHALF', '')
            sfla = row.get('SFLA', '')

            record_mismatches = []
            fields_compared = []

            for mapping in cols_to_compare_mapping:
                mls_col = mapping['mls_col']
                cama_col = mapping['cama_col']
                if mls_col not in merged_df.columns or cama_col not in merged_df.columns:
                    continue
                mls_val = row.get(mls_col)
                cama_val = row.get(cama_col)
                mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')
                if mls_is_blank or cama_is_blank:
                    continue
                fields_compared.append(mls_col)
                if SKIP_ZERO_VALUES:
                    try:
                        mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
                        if (pd.notna(mls_numeric) and mls_numeric == 0) and (pd.notna(cama_numeric) and cama_numeric == 0):
                            continue
                    except:
                        pass
                if not values_equal(mls_val, cama_val):
                    record_mismatches.append({
                        'Parcel_ID': record_id,
                        'NOPAR': nopar,
                        'ADDITIONAL_PARCELS': additional_parcels,
                        'Listing_Number': listing_num,
                        'SALEKEY': salekey,
                        'Address': address,
                        'City': city,
                        'State': state,
                        'Zip': zip_code,
                        'Zone': zone,
                        'YRBLT': yrblt,
                        'EFFYR': effyr,
                        'GRADE': grade,
                        'CDU': cdu,
                        'FIXBATH': fixbath,
                        'FIXHALF': fixhalf,
                        'SFLA': sfla,
                        'Public_Remarks': public_remarks,
                        'Field_MLS': mls_col,
                        'Field_CAMA': cama_col,
                        'MLS_Value': mls_val,
                        'CAMA_Value': cama_val,
                        'Difference': calculate_difference(mls_val, cama_val),
                        'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                        'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                    })

            if cols_to_compare_sum:
                for mapping in cols_to_compare_sum:
                    mls_col = mapping['mls_col']
                    cama_cols = mapping['cama_cols']
                    if mls_col not in merged_df.columns:
                        continue
                    missing_cama_cols = [col for col in cama_cols if col not in merged_df.columns]
                    if missing_cama_cols:
                        continue
                    mls_val = row.get(mls_col)
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    if mls_is_blank:
                        continue
                    cama_sum = 0
                    all_cama_blank = True
                    for col in cama_cols:
                        val = row.get(col)
                        if pd.notna(val):
                            all_cama_blank = False
                            try:
                                cama_sum += pd.to_numeric(val, errors='coerce')
                            except:
                                pass
                    if all_cama_blank:
                        continue
                    fields_compared.append(mls_col)
                    if SKIP_ZERO_VALUES:
                        try:
                            mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                            if (pd.notna(mls_numeric) and mls_numeric == 0) and (cama_sum == 0):
                                continue
                        except:
                            pass
                    if not values_equal(mls_val, cama_sum):
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'NOPAR': nopar,
                            'ADDITIONAL_PARCELS': additional_parcels,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Zone': zone,
                            'YRBLT': yrblt,
                            'EFFYR': effyr,
                            'GRADE': grade,
                            'CDU': cdu,
                            'FIXBATH': fixbath,
                            'FIXHALF': fixhalf,
                            'SFLA': sfla,
                            'Public_Remarks': public_remarks,
                            'Field_MLS': mls_col,
                            'Field_CAMA': f"SUM({', '.join(cama_cols)})",
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_sum,
                            'Difference': calculate_difference(mls_val, cama_sum),
                            'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                            'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                        })

            if cols_to_compare_categorical:
                for mapping in cols_to_compare_categorical:
                    mls_col = mapping['mls_col']
                    cama_col = mapping['cama_col']
                    if mls_col not in merged_df.columns or cama_col not in merged_df.columns:
                        continue
                    mls_val = row.get(mls_col)
                    cama_val = row.get(cama_col)
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')
                    if mls_is_blank or cama_is_blank:
                        continue
                    fields_compared.append(mls_col)
                    is_match = categorical_match(mls_val, cama_val, mapping)
                    check_text = mapping.get('mls_check_contains', '')
                    case_sensitive = mapping.get('case_sensitive', False)
                    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''
                    if not case_sensitive:
                        mls_str = mls_str.lower()
                        check_text_lower = check_text.lower()
                    else:
                        check_text_lower = check_text
                    text_found = check_text_lower in mls_str
                    expected_cama = mapping.get('cama_expected_if_true') if text_found else mapping.get('cama_expected_if_false')
                    if not is_match:
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'NOPAR': nopar,
                            'ADDITIONAL_PARCELS': additional_parcels,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Zone': zone,
                            'YRBLT': yrblt,
                            'EFFYR': effyr,
                            'GRADE': grade,
                            'CDU': cdu,
                            'FIXBATH': fixbath,
                            'FIXHALF': fixhalf,
                            'SFLA': sfla,
                            'Public_Remarks': public_remarks,
                            'Field_MLS': mls_col,
                            'Field_CAMA': cama_col,
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_val,
                            'Expected_CAMA_Value': expected_cama,
                            'Match_Rule': f"If '{check_text}' in {mls_col}, then {cama_col} should be {mapping.get('cama_expected_if_true')}, else {mapping.get('cama_expected_if_false')}",
                            'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                            'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                        })

            if not record_mismatches and fields_compared:
                perfect_matches.append({
                    'Parcel_ID': record_id,
                    'NOPAR': nopar,
                    'ADDITIONAL_PARCELS': additional_parcels,
                    'Listing_Number': listing_num,
                    'SALEKEY': salekey,
                    'Address': address,
                    'City': city,
                    'State': state,
                    'Zip': zip_code,
                    'Zone': zone,
                    'YRBLT': yrblt,
                    'EFFYR': effyr,
                    'GRADE': grade,
                    'CDU': cdu,
                    'FIXBATH': fixbath,
                    'FIXHALF': fixhalf,
                    'SFLA': sfla,
                    'Public_Remarks': public_remarks,
                    'Fields_Compared': len(fields_compared),
                    'Fields_List': ', '.join(fields_compared),
                    'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                    'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                })

            value_mismatches.extend(record_mismatches)

    return (pd.DataFrame(missing_in_cama), pd.DataFrame(missing_in_mls),
            pd.DataFrame(value_mismatches), matched_df, pd.DataFrame(perfect_matches))

def create_excel_with_hyperlinks(df, sheet_name='Sheet1'):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    wb = load_workbook(output)
    ws = wb[sheet_name]

    if 'Parcel_ID' in df.columns and 'Parcel_URL' in df.columns:
        parcel_col_idx = list(df.columns).index('Parcel_ID') + 1
        url_col_idx = list(df.columns).index('Parcel_URL') + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=parcel_col_idx)
            url = ws.cell(row=row_idx, column=url_col_idx).value
            if url and str(url).strip() and str(url) != 'nan':
                cell.hyperlink = url
                cell.style = 'Hyperlink'

    if 'Address' in df.columns and 'Zillow_URL' in df.columns:
        address_col_idx = list(df.columns).index('Address') + 1
        zillow_col_idx = list(df.columns).index('Zillow_URL') + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=address_col_idx)
            url = ws.cell(row=row_idx, column=zillow_col_idx).value
            if url and str(url).strip() and str(url) != 'nan':
                cell.hyperlink = url
                cell.style = 'Hyperlink'

    if 'Parcel_URL' in df.columns:
        url_col_idx = list(df.columns).index('Parcel_URL') + 1
        ws.delete_cols(url_col_idx)

    if 'Zillow_URL' in df.columns:
        remaining_cols = [col for col in df.columns if col not in ['Parcel_URL']]
        if 'Zillow_URL' in remaining_cols:
            zillow_col_idx = remaining_cols.index('Zillow_URL') + 1
            ws.delete_cols(zillow_col_idx)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()

def create_zip_with_all_reports(df_missing_cama, df_missing_mls, df_value_mismatches,
                                 df_perfect_matches, city_comparison_df=None):
    import zipfile
    timestamp = datetime.now().strftime("%Y-%m-%d")
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        if not df_missing_cama.empty:
            zip_file.writestr(f"missing_in_CAMA_{timestamp}.xlsx",
                              create_excel_with_hyperlinks(df_missing_cama, 'Missing in CAMA'))
        if not df_missing_mls.empty:
            zip_file.writestr(f"missing_in_MLS_{timestamp}.xlsx",
                              create_excel_with_hyperlinks(df_missing_mls, 'Missing in MLS'))
        if not df_value_mismatches.empty:
            zip_file.writestr(f"value_mismatches_{timestamp}.xlsx",
                              create_excel_with_hyperlinks(df_value_mismatches, 'Value Mismatches'))
        if not df_perfect_matches.empty:
            zip_file.writestr(f"perfect_matches_{timestamp}.xlsx",
                              create_excel_with_hyperlinks(df_perfect_matches, 'Perfect Matches'))
        if city_comparison_df is not None and not city_comparison_df.empty:
            zip_file.writestr(f"city_match_statistics_{timestamp}.csv",
                              city_comparison_df.to_csv(index=False))
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

# --- Streamlit App ---

st.title("📊 MLS vs CAMA Data Comparison Tool")
st.markdown("Compare MLS and CAMA property data to identify discrepancies and perfect matches.")

with st.sidebar:
    st.header("⚙️ Configuration")

    st.subheader("WindowId Setup")
    st.markdown("""
    **How to get WindowId:**
    1. Go to [Stark County iasWorld](https://iasworld.starkcountyohio.gov/iasworld/)
    2. Log in and search for any property
    3. Copy the `windowId` value from the URL

    Example: `...windowId=638981240146803746&...`
    """)

    window_id = st.text_input(
        "Enter WindowId",
        value="638981240146803746",
        help="Used to generate clickable links to property records"
    )

    st.divider()

    st.subheader("Comparison Settings")
    tolerance = st.number_input(
        "Numeric Tolerance",
        value=0.01,
        format="%.4f",
        help="Absolute tolerance for numeric comparisons"
    )

    skip_zeros = st.checkbox(
        "Skip Zero Values",
        value=True,
        help="If enabled, skips comparison only when BOTH values are zero"
    )

NUMERIC_TOLERANCE = tolerance
SKIP_ZERO_VALUES = skip_zeros

st.header("📁 Upload Data Files")

col1, col2 = st.columns(2)

with col1:
    mls_file = st.file_uploader(
        "Upload MLS Data (Excel)",
        type=['xlsx', 'xls'],
        help="Blanks are automatically filled with 0 — no CTRL+H needed"
    )

with col2:
    cama_file = st.file_uploader(
        "Upload CAMA Data (Excel)",
        type=['xlsx', 'xls'],
        help="Upload your CAMA Excel file exported from SQL Developer"
    )

if mls_file and cama_file:

    with st.spinner("Loading data files..."):
        try:
            # Auto-fill MLS blanks with 0 — replaces manual CTRL+H step
            df_mls, blank_count = load_mls_file(mls_file)
            df_cama = pd.read_excel(cama_file)
            st.success(
                f"✅ Files loaded! MLS: **{len(df_mls):,} records** "
                f"({blank_count:,} blank cells auto-filled with 0). "
                f"CAMA: **{len(df_cama):,} records**."
            )
        except Exception as e:
            st.error(f"Error loading files: {e}")
            st.stop()

    st.header("📊 Data Summary")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("MLS Records", len(df_mls))
    with col2:
        st.metric("CAMA Records", len(df_cama))
    with col3:
        st.metric("Numeric Tolerance", f"{NUMERIC_TOLERANCE}")

    col1, col2 = st.columns([3, 1])
    with col1:
        run_button = st.button("🔍 Run Comparison", type="primary", use_container_width=True)
    with col2:
        if st.session_state.get('comparison_complete', False):
            if st.button("🔄 Clear Results", use_container_width=True):
                for key in ['df_missing_cama', 'df_missing_mls', 'df_value_mismatches',
                            'matched_df', 'df_perfect_matches', 'comparison_complete', 'city_comparison']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()

    if run_button:
        with st.spinner("Comparing data... This may take a moment."):
            df_missing_cama, df_missing_mls, df_value_mismatches, matched_df, df_perfect_matches = \
                compare_data_enhanced(
                    df_mls, df_cama,
                    UNIQUE_ID_COLUMN,
                    COLUMNS_TO_COMPARE,
                    cols_to_compare_sum=COLUMNS_TO_COMPARE_SUM,
                    cols_to_compare_categorical=COLUMNS_TO_COMPARE_CATEGORICAL,
                    window_id=window_id
                )
            st.session_state['df_missing_cama'] = df_missing_cama
            st.session_state['df_missing_mls'] = df_missing_mls
            st.session_state['df_value_mismatches'] = df_value_mismatches
            st.session_state['matched_df'] = matched_df
            st.session_state['df_perfect_matches'] = df_perfect_matches
            st.session_state['comparison_complete'] = True

    if st.session_state.get('comparison_complete', False):
        df_missing_cama = st.session_state['df_missing_cama']
        df_missing_mls = st.session_state['df_missing_mls']
        df_value_mismatches = st.session_state['df_value_mismatches']
        matched_df = st.session_state['matched_df']
        df_perfect_matches = st.session_state['df_perfect_matches']

        st.success("✅ Comparison complete. Scroll down to download reports or generate mass update files.")

        st.header("📈 Results Summary")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("✅ Matched Records", len(matched_df))
        with col2:
            st.metric("❌ Missing in CAMA", len(df_missing_cama))
        with col3:
            st.metric("❌ Missing in MLS", len(df_missing_mls))
        with col4:
            st.metric("⚠️ Value Mismatches", len(df_value_mismatches))

        col1, col2 = st.columns(2)
        with col1:
            st.metric("✅ Perfect Matches", len(df_perfect_matches))
        with col2:
            if not df_value_mismatches.empty:
                st.metric("📊 Fields with Mismatches", df_value_mismatches['Field_MLS'].nunique())

        # City match stats
        st.header("📊 CAMA Parcel Match Statistics")
        total_cama_parcels = len(df_cama)
        matched_parcels = len(matched_df)
        match_rate = (matched_parcels / total_cama_parcels * 100) if total_cama_parcels > 0 else 0

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total CAMA Parcels", f"{total_cama_parcels:,}")
        with col2:
            st.metric("Found in MLS", f"{matched_parcels:,}")
        with col3:
            st.metric("Match Rate", f"{match_rate:.2f}%")

        st.subheader("Match Rate by City")
        cama_city_col = 'CITYNAME' if 'CITYNAME' in df_cama.columns else ('City' if 'City' in df_cama.columns else None)

        if cama_city_col and not matched_df.empty:
            cama_id_col = UNIQUE_ID_COLUMN.get('cama_col')
            if cama_city_col in matched_df.columns:
                cama_cities = df_cama.groupby(cama_city_col)[cama_id_col].count().reset_index()
                cama_cities.columns = ['City', 'Total_CAMA_Parcels']
                matched_cities = matched_df.groupby(cama_city_col)[cama_id_col].count().reset_index()
                matched_cities.columns = ['City', 'Matched_Parcels']
                city_comparison = pd.merge(cama_cities, matched_cities, on='City', how='left')
                city_comparison['Matched_Parcels'] = city_comparison['Matched_Parcels'].fillna(0).astype(int)
                city_comparison['Match_Rate'] = (city_comparison['Matched_Parcels'] / city_comparison['Total_CAMA_Parcels'] * 100).round(2)
                city_comparison['Not_Matched'] = city_comparison['Total_CAMA_Parcels'] - city_comparison['Matched_Parcels']
                city_comparison = city_comparison.sort_values('Total_CAMA_Parcels', ascending=False)
                st.session_state['city_comparison'] = city_comparison
                st.dataframe(
                    city_comparison[['City', 'Total_CAMA_Parcels', 'Matched_Parcels', 'Not_Matched', 'Match_Rate']],
                    use_container_width=True, hide_index=True
                )
                st.download_button(
                    label="📥 Download City Statistics (CSV)",
                    data=city_comparison.to_csv(index=False),
                    file_name=f"city_match_statistics_{datetime.now().strftime('%Y-%m-%d')}.csv",
                    mime="text/csv"
                )
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Top 10 Cities by Total CAMA Parcels**")
                    st.bar_chart(city_comparison.head(10).set_index('City')[['Matched_Parcels', 'Not_Matched']])
                with col2:
                    st.markdown("**Match Rate by City (Top 10)**")
                    st.bar_chart(city_comparison.head(10)[['City', 'Match_Rate']].set_index('City'))
            else:
                st.info("ℹ️ City column from CAMA not available in matched records")
        else:
            st.info("ℹ️ City information not available in the data")

        if not df_value_mismatches.empty:
            st.subheader("📊 Mismatches by Field")
            st.bar_chart(df_value_mismatches['Field_MLS'].value_counts())

        # Data preview tabs
        st.header("📋 Data Preview")
        tab1, tab2, tab3, tab4 = st.tabs(["Missing in CAMA", "Missing in MLS", "Value Mismatches", "Perfect Matches"])
        with tab1:
            if not df_missing_cama.empty:
                st.dataframe(df_missing_cama, use_container_width=True)
            else:
                st.info("No records missing in CAMA")
        with tab2:
            if not df_missing_mls.empty:
                st.dataframe(df_missing_mls, use_container_width=True)
            else:
                st.info("No records missing in MLS")
        with tab3:
            if not df_value_mismatches.empty:
                st.dataframe(df_value_mismatches, use_container_width=True)
            else:
                st.info("No value mismatches found")
        with tab4:
            if not df_perfect_matches.empty:
                st.dataframe(df_perfect_matches, use_container_width=True)
            else:
                st.info("No perfect matches found")

        # Download section
        st.header("📥 Download Reports")

        st.markdown("### 📦 Download All Reports")
        city_comp = st.session_state.get('city_comparison', None)
        timestamp = datetime.now().strftime("%Y-%m-%d")
        st.download_button(
            label="📦 Download All Reports (ZIP)",
            data=create_zip_with_all_reports(df_missing_cama, df_missing_mls, df_value_mismatches, df_perfect_matches, city_comp),
            file_name=f"MLS_CAMA_Comparison_All_Reports_{timestamp}.zip",
            mime="application/zip",
            use_container_width=True
        )

        st.markdown("### 📄 Download Individual Reports")
        col1, col2 = st.columns(2)
        with col1:
            if not df_missing_cama.empty:
                st.download_button(
                    label="📄 Download Missing in CAMA",
                    data=create_excel_with_hyperlinks(df_missing_cama, 'Missing in CAMA'),
                    file_name=f"missing_in_CAMA_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            if not df_value_mismatches.empty:
                st.download_button(
                    label="⚠️ Download Value Mismatches",
                    data=create_excel_with_hyperlinks(df_value_mismatches, 'Value Mismatches'),
                    file_name=f"value_mismatches_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        with col2:
            if not df_missing_mls.empty:
                st.download_button(
                    label="📄 Download Missing in MLS",
                    data=create_excel_with_hyperlinks(df_missing_mls, 'Missing in MLS'),
                    file_name=f"missing_in_MLS_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            if not df_perfect_matches.empty:
                st.download_button(
                    label="✅ Download Perfect Matches",
                    data=create_excel_with_hyperlinks(df_perfect_matches, 'Perfect Matches'),
                    file_name=f"perfect_matches_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # ── MLS Search Batches ──────────────────────────────────────────
        if not df_missing_mls.empty:
            st.markdown("### 🔍 MLS Search Batches (Missing in MLS)")
            st.markdown(
                f"**{len(df_missing_mls)} parcels** not found in MLS. "
                "Download the file below — parcel numbers are pre-formatted to 8 digits with leading zeros "
                "and split into batches of 35, one tab per batch. "
                "Open a tab, copy column A, and paste directly into the MLS parcel search."
            )
            batch_excel, num_batches = create_mls_search_batches_excel(df_missing_mls, batch_size=35)
            timestamp_file = datetime.now().strftime("%m%d%y")
            st.download_button(
                label=f"🔍 Download MLS Search Batches ({num_batches} {'batch' if num_batches == 1 else 'batches'} of 35)",
                data=batch_excel,
                file_name=f"MLS_Search_Batches_{timestamp_file}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            with st.expander("👁️ Preview first batch (comma-separated)"):
                preview_parcels = []
                for p in df_missing_mls['Parcel_ID'].head(35).tolist():
                    try:
                        preview_parcels.append(str(int(float(str(p)))).zfill(8))
                    except:
                        preview_parcels.append(str(p).zfill(8))
                st.code(','.join(preview_parcels), language=None)

        # ── CAMA Mass Updates ───────────────────────────────────────────
        st.header("🔄 CAMA Mass Updates")
        st.markdown("Generate mass update files for CAMA system from perfect matches and value mismatches.")

        if not df_perfect_matches.empty or not df_value_mismatches.empty:
            with st.expander("⚙️ Generate CAMA Mass Update Files", expanded=False):
                st.markdown("""
                Generates two files for mass updating the CAMA system:
                1. **SALETAB_MassUpdate** — Updates the SALE tab
                2. **MassEntrance** — Updates the ENTRANCE tab
                """)
                col1, col2 = st.columns(2)
                with col1:
                    user_initials = st.text_input("Your Initials (e.g., JMJ)", value="JMJ")
                with col2:
                    user_full_name = st.text_input("Your Full Name (e.g., Jason Jeffries)", value="Jason Jeffries")

                if st.button("🔄 Generate Mass Update Files", type="primary"):
                    combined_df = pd.concat([df_perfect_matches, df_value_mismatches], ignore_index=True)
                    if combined_df.empty:
                        st.error("No data available to generate mass update files")
                    else:
                        st.info(f"📋 Combined data has {len(combined_df)} records")
                        if 'Listing_Number' not in combined_df.columns:
                            st.error("❌ 'Listing_Number' column not found!")
                            st.write("Available columns:", combined_df.columns.tolist())
                        else:
                            non_null = combined_df['Listing_Number'].notna().sum()
                            st.info(f"✔ Listing_Number found. Non-null values: {non_null}/{len(combined_df)}")
                            with st.expander("📊 Sample Data Preview"):
                                st.dataframe(combined_df[['Parcel_ID', 'Listing_Number']].head(5))

                        saletab_data, entrance_data = generate_mass_update_files(combined_df, user_initials, user_full_name)
                        st.success(f"✅ Generated mass update files for {len(combined_df)} records")

                        col1, col2 = st.columns(2)
                        timestamp_file = datetime.now().strftime("%m%d%y")
                        with col1:
                            st.download_button(
                                label="📥 Download SALETAB_MassUpdate",
                                data=saletab_data,
                                file_name=f"SALETAB_MassUpdate_{timestamp_file}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        with col2:
                            st.download_button(
                                label="📥 Download MassEntrance",
                                data=entrance_data,
                                file_name=f"MassEntrance{timestamp_file}.csv",
                                mime="text/csv"
                            )
        else:
            st.info("ℹ️ No data available for mass updates. Run comparison first.")

else:
    st.info("👆 Please upload both MLS and CAMA data files to begin.")

    with st.expander("ℹ️ Expected Data Format"):
        st.markdown("""
        ### MLS Data Expected Columns:
        - `Parcel Number` (unique identifier)
        - `Above Grade Finished Area`, `Bedrooms Total`, `Bathrooms Full`, `Bathrooms Half`
        - `Below Grade Finished Area`, `Cooling`
        - `Address`, `City`, `State or Province`, `Postal Code`

        **Note:** Blank cells are automatically filled with 0 on upload — no manual CTRL+H needed.

        ### CAMA Data Expected Columns:
        - `PARID` (unique identifier)
        - `NOPAR`, `CITYNAME` (or `City`)
        - `SFLA`, `RMBED`, `FIXBATH`, `FIXHALF`
        - `RECROMAREA`, `FINBSMTAREA`, `UFEATAREA`
        - `HEAT`, `SALEKEY`
        """)

st.divider()
st.caption("MLS vs CAMA Comparison Tool | Built with Streamlit")
